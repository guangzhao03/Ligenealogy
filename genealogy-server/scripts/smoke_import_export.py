"""R7 Excel import/export smoke test."""
from __future__ import annotations

import json
import sys
import time
import uuid
import urllib.parse
import urllib.request
from io import BytesIO

from openpyxl import Workbook, load_workbook

BASE = "http://127.0.0.1:8000"

PERSON_HEADERS = [
    "姓名",
    "性别",
    "世代",
    "出生日期",
    "去世日期",
    "籍贯",
    "简介",
    "备注",
    "是否在世",
]
RELATION_HEADERS = ["from姓名", "to姓名", "关系类型"]


def api_json(method: str, path: str, body: dict | None = None, token: str | None = None) -> dict:
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(f"{BASE}{path}", data=data, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode("utf-8"))


def build_xlsx(headers: list[str], rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def upload_excel(token: str, path: str, family_id: int, filename: str, content: bytes) -> dict:
    boundary = f"----Boundary{uuid.uuid4().hex}"
    parts: list[bytes] = []

    def add_field(name: str, value: str) -> None:
        parts.append(
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
            f"{value}\r\n".encode("utf-8")
        )

    add_field("family_id", str(family_id))
    parts.append(
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
            f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8")
        + content
        + b"\r\n"
    )
    parts.append(f"--{boundary}--\r\n".encode("utf-8"))

    req = urllib.request.Request(
        f"{BASE}{path}",
        data=b"".join(parts),
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Authorization": f"Bearer {token}",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode("utf-8"))


def download_export(token: str, path: str) -> bytes:
    req = urllib.request.Request(
        f"{BASE}{path}",
        headers={"Authorization": f"Bearer {token}"},
        method="GET",
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        return resp.read()


def main() -> None:
    username = f"excel_user_{int(time.time())}"
    api_json("POST", "/api/auth/register", {"username": username, "password": "123456"})
    token = api_json("POST", "/api/auth/login", {"username": username, "password": "123456"})[
        "data"
    ]["access_token"]

    family_id = api_json("POST", "/api/families", {"name": "赵氏家族"}, token=token)["data"]["id"]

    persons_xlsx = build_xlsx(
        PERSON_HEADERS,
        [
            ["赵大", 1, 1, "1950-01-01", None, "福建", "始祖", "备注1", 1],
            ["赵二", 1, 2, "1980-05-05", None, "福建", "", "", 1],
        ],
    )
    imported_persons = upload_excel(
        token, "/api/import/persons", family_id, "persons.xlsx", persons_xlsx
    )
    assert imported_persons["code"] == 0, imported_persons
    assert imported_persons["data"]["success_count"] == 2
    print("[OK] import persons")

    export_persons_bytes = download_export(
        token, f"/api/export/persons?{urllib.parse.urlencode({'family_id': family_id})}"
    )
    workbook = load_workbook(BytesIO(export_persons_bytes), data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == tuple(PERSON_HEADERS)
    assert len(rows) == 3
    print("[OK] export persons")

    relations_xlsx = build_xlsx(
        RELATION_HEADERS,
        [["赵大", "赵二", "parent"]],
    )
    imported_relations = upload_excel(
        token, "/api/import/relations", family_id, "relations.xlsx", relations_xlsx
    )
    assert imported_relations["code"] == 0, imported_relations
    assert imported_relations["data"]["success_count"] == 1
    print("[OK] import relations")

    export_relations_bytes = download_export(
        token, f"/api/export/relations?{urllib.parse.urlencode({'family_id': family_id})}"
    )
    rel_book = load_workbook(BytesIO(export_relations_bytes), data_only=True)
    rel_rows = list(rel_book.active.iter_rows(values_only=True))
    assert rel_rows[1][0] == "赵大" and rel_rows[1][1] == "赵二"
    assert rel_rows[1][2] == "parent"
    print("[OK] export relations")

    bad_xlsx = build_xlsx(PERSON_HEADERS, [["", 1, 1, None, None, None, None, None, 1]])
    bad_result = upload_excel(token, "/api/import/persons", family_id, "bad.xlsx", bad_xlsx)
    assert bad_result["data"]["success_count"] == 0
    assert bad_result["data"]["errors"]
    print("[OK] invalid row rejected")

    print("\nALL PASSED")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\nFAILED: {exc}", file=sys.stderr)
        sys.exit(1)
