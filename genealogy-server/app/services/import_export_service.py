from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.person import Person
from app.models.person_relation import PersonRelation
from app.models.user import User
from app.schemas.import_export import ImportErrorItem, ImportResult
from app.schemas.relation import RelationCreate
from app.schemas.relation_type import RelationType
from app.services import family_service
from app.utils.exceptions import BadRequestException

PERSON_HEADERS = [
    "姓名",
    "小名",
    "性别",
    "辈分",
    "出生年份",
    "出生日期",
    "去世日期",
    "籍贯",
    "简介",
    "备注",
    "是否在世",
]
RELATION_HEADERS = ["from姓名", "to姓名", "关系类型"]

GENDER_MAP = {"男": 1, "女": 2, "未知": 0, "0": 0, "1": 1, "2": 2}
ALIVE_MAP = {"是": 1, "否": 0, "1": 1, "0": 0, "在世": 1, "已故": 0}


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def _parse_date(value: Any) -> date | None:
    value = _cell_value(value)
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    raise ValueError("日期格式应为 YYYY-MM-DD")


def _parse_int(value: Any, field_name: str) -> int | None:
    value = _cell_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        return int(value)
    raise ValueError(f"{field_name}应为整数")


def _parse_gender(value: Any) -> int:
    value = _cell_value(value)
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    key = str(value)
    if key in GENDER_MAP:
        return GENDER_MAP[key]
    raise ValueError("性别应为 0/1/2 或 男/女/未知")


def _parse_alive(value: Any) -> int:
    value = _cell_value(value)
    if value is None:
        return 1
    if isinstance(value, (int, float)):
        return int(value)
    key = str(value)
    if key in ALIVE_MAP:
        return ALIVE_MAP[key]
    raise ValueError("是否在世应为 0/1 或 是/否")


def _read_sheet_rows(file_bytes: bytes) -> list[list[Any]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        rows.append(list(row))
    if not rows:
        raise BadRequestException("Excel 文件为空")
    return rows


def _validate_headers(actual: list[Any], expected: list[str]) -> None:
    headers = [str(cell).strip() if cell is not None else "" for cell in actual]
    if headers != expected:
        raise BadRequestException(f"表头不正确，应为: {', '.join(expected)}")


async def import_persons(
    db: Session, current_user: User, family_id: int, file: UploadFile
) -> ImportResult:
    family_service.get_family(db, family_id, current_user)
    rows = _read_sheet_rows(await file.read())
    _validate_headers(rows[0], PERSON_HEADERS)

    parsed: list[Person] = []
    errors: list[ImportErrorItem] = []

    for index, row in enumerate(rows[1:], start=2):
        try:
            name = _cell_value(row[0])
            nickname = _cell_value(row[1])
            if not name:
                raise ValueError("姓名不能为空")
            if not nickname:
                raise ValueError("小名不能为空")
            generation = _parse_int(row[3], "辈分")
            birth_year = _parse_int(row[4], "出生年份")
            if generation is None:
                raise ValueError("辈分不能为空")
            if birth_year is None:
                raise ValueError("出生年份不能为空")
            person = Person(
                family_id=family_id,
                name=str(name),
                nickname=str(nickname),
                gender=_parse_gender(row[2]),
                generation=generation,
                birth_year=birth_year,
                birth_date=_parse_date(row[5]),
                death_date=_parse_date(row[6]),
                birthplace=_cell_value(row[7]),
                biography=_cell_value(row[8]),
                remark=_cell_value(row[9]),
                is_alive=_parse_alive(row[10]),
            )
            parsed.append(person)
        except Exception as exc:
            errors.append(ImportErrorItem(row=index, message=str(exc)))

    if errors:
        return ImportResult(success_count=0, errors=errors)

    for person in parsed:
        db.add(person)
    db.commit()
    return ImportResult(success_count=len(parsed), errors=[])


async def import_relations(
    db: Session, current_user: User, family_id: int, file: UploadFile
) -> ImportResult:
    family_service.get_family(db, family_id, current_user)
    rows = _read_sheet_rows(await file.read())
    _validate_headers(rows[0], RELATION_HEADERS)

    persons = db.scalars(select(Person).where(Person.family_id == family_id)).all()
    name_map: dict[str, list[Person]] = {}
    for person in persons:
        name_map.setdefault(person.name, []).append(person)

    parsed: list[RelationCreate] = []
    errors: list[ImportErrorItem] = []

    for index, row in enumerate(rows[1:], start=2):
        try:
            from_name = _cell_value(row[0])
            to_name = _cell_value(row[1])
            relation_type = _cell_value(row[2])
            if not from_name or not to_name or not relation_type:
                raise ValueError("from姓名、to姓名、关系类型均不能为空")

            from_matches = name_map.get(str(from_name), [])
            to_matches = name_map.get(str(to_name), [])
            if len(from_matches) != 1:
                raise ValueError(f"from姓名「{from_name}」未唯一匹配到人物")
            if len(to_matches) != 1:
                raise ValueError(f"to姓名「{to_name}」未唯一匹配到人物")

            if relation_type not in {RelationType.parent.value, RelationType.spouse.value}:
                raise ValueError("关系类型仅支持 parent 或 spouse")

            parsed.append(
                RelationCreate(
                    family_id=family_id,
                    from_person_id=from_matches[0].id,
                    to_person_id=to_matches[0].id,
                    relation_type=RelationType(relation_type),
                )
            )
        except Exception as exc:
            errors.append(ImportErrorItem(row=index, message=str(exc)))

    if errors:
        return ImportResult(success_count=0, errors=errors)

    seen: set[tuple[int, int, str]] = set()
    for item in parsed:
        key = (item.from_person_id, item.to_person_id, item.relation_type.value)
        if item.relation_type == RelationType.spouse:
            key = (
                min(item.from_person_id, item.to_person_id),
                max(item.from_person_id, item.to_person_id),
                item.relation_type.value,
            )
        if key in seen:
            return ImportResult(
                success_count=0,
                errors=[ImportErrorItem(row=0, message="导入文件中存在重复关系")],
            )
        seen.add(key)

    try:
        for item in parsed:
            from_id = item.from_person_id
            to_id = item.to_person_id
            relation_type = item.relation_type.value
            if relation_type == RelationType.spouse.value:
                from_id, to_id = min(from_id, to_id), max(from_id, to_id)
            db.add(
                PersonRelation(
                    family_id=family_id,
                    from_person_id=from_id,
                    to_person_id=to_id,
                    relation_type=relation_type,
                )
            )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return ImportResult(success_count=len(parsed), errors=[])


def _workbook_response(workbook: Workbook, filename: str) -> tuple[BytesIO, str]:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer, filename


def export_persons(db: Session, current_user: User, family_id: int) -> tuple[BytesIO, str]:
    family_service.get_family(db, family_id, current_user)
    persons = db.scalars(
        select(Person).where(Person.family_id == family_id).order_by(Person.id.asc())
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "persons"
    sheet.append(PERSON_HEADERS)
    for person in persons:
        sheet.append(
            [
                person.name,
                person.nickname,
                person.gender,
                person.generation,
                person.birth_year,
                person.birth_date,
                person.death_date,
                person.birthplace,
                person.biography,
                person.remark,
                person.is_alive,
            ]
        )
    return _workbook_response(workbook, f"persons_{family_id}.xlsx")


def export_relations(db: Session, current_user: User, family_id: int) -> tuple[BytesIO, str]:
    family_service.get_family(db, family_id, current_user)
    persons = db.scalars(select(Person).where(Person.family_id == family_id)).all()
    person_name_map = {person.id: person.name for person in persons}
    relations = db.scalars(
        select(PersonRelation)
        .where(PersonRelation.family_id == family_id)
        .order_by(PersonRelation.id.asc())
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "relations"
    sheet.append(RELATION_HEADERS)
    for relation in relations:
        sheet.append(
            [
                person_name_map.get(relation.from_person_id, ""),
                person_name_map.get(relation.to_person_id, ""),
                relation.relation_type,
            ]
        )
    return _workbook_response(workbook, f"relations_{family_id}.xlsx")
